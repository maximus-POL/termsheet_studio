from __future__ import annotations

from pathlib import Path
from typing import Any

from schema import flatten_compact_product, is_compact_product
from template_mapping import TEMPLATE_CELL_MAPPING

MAPPING_SHEET_NAME = "Field Mapping"
SUPPORTED_EXCEL_TEMPLATE_SUFFIXES = {".xlsx", ".xlsm"}


class ExcelTemplateError(RuntimeError):
    pass


def write_product_excel(product_data: dict[str, Any], template_path: Path, output_path: Path) -> Path:
    if not template_path.exists():
        raise ExcelTemplateError(f"Template not found: {template_path}")

    if template_path.suffix.lower() not in SUPPORTED_EXCEL_TEMPLATE_SUFFIXES:
        raise ExcelTemplateError(f"Unsupported Excel template type: {template_path.suffix}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelTemplateError(
            "openpyxl is not installed. Run: pip install -r requirements.txt"
        ) from exc

    output_path = output_path.with_suffix(template_path.suffix.lower())
    workbook = load_workbook(template_path, keep_vba=is_macro_enabled_workbook(template_path))
    cell_mapping = load_template_cell_mapping(workbook)
    fields = resolve_product_fields(product_data)
    if not isinstance(fields, dict):
        raise ExcelTemplateError("product_data must contain a 'fields' dictionary")

    for field_name, target in cell_mapping.items():
        value = fields.get(field_name)
        if value in (None, ""):
            continue

        worksheet, cell = resolve_target(workbook, target)
        worksheet[cell] = value

    remove_mapping_sheet(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def resolve_product_fields(product_data: dict[str, Any]) -> dict[str, Any]:
    fields = product_data.get("fields", {})
    if isinstance(fields, dict) and is_compact_product(fields):
        return flatten_compact_product(fields)
    return fields


def is_macro_enabled_workbook(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def load_template_cell_mapping(workbook: Any) -> dict[str, str | dict[str, str | None]]:
    if MAPPING_SHEET_NAME not in workbook.sheetnames:
        return TEMPLATE_CELL_MAPPING

    worksheet = workbook[MAPPING_SHEET_NAME]
    headers = get_mapping_headers(worksheet)
    field_col = headers.get("field_name")
    cell_col = headers.get("cell")
    sheet_col = headers.get("sheet")

    if not field_col or not cell_col:
        raise ExcelTemplateError(
            f"'{MAPPING_SHEET_NAME}' sheet must have 'field_name' and 'cell' headers"
        )

    mapping: dict[str, dict[str, str | None]] = {}
    for row_number in range(2, worksheet.max_row + 1):
        field_name = normalize_cell_text(worksheet.cell(row_number, field_col).value)
        if not field_name:
            continue

        cell = normalize_cell_text(worksheet.cell(row_number, cell_col).value)
        if not cell:
            raise ExcelTemplateError(
                f"'{MAPPING_SHEET_NAME}' row {row_number} has field_name "
                f"'{field_name}' but no target cell"
            )

        sheet_name = (
            normalize_cell_text(worksheet.cell(row_number, sheet_col).value)
            if sheet_col
            else ""
        )
        mapping[field_name] = {"sheet": sheet_name or None, "cell": cell}

    if not mapping:
        raise ExcelTemplateError(f"'{MAPPING_SHEET_NAME}' sheet does not define any mappings")

    return mapping


def get_mapping_headers(worksheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in worksheet[1]:
        header = normalize_header(cell.value)
        if header:
            headers[header] = cell.column
    return headers


def normalize_header(value: Any) -> str:
    return normalize_cell_text(value).lower().replace(" ", "_")


def normalize_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def resolve_target(workbook: Any, target: str | dict[str, str | None]) -> tuple[Any, str]:
    if isinstance(target, str):
        return workbook.active, target

    sheet_name = target.get("sheet")
    cell = target.get("cell")
    if not cell:
        raise ExcelTemplateError(f"Invalid template mapping target: {target}")

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ExcelTemplateError(f"Template sheet not found: {sheet_name}")
        return workbook[sheet_name], cell

    return workbook.active, cell


def remove_mapping_sheet(workbook: Any) -> None:
    if MAPPING_SHEET_NAME not in workbook.sheetnames:
        return

    if len(workbook.sheetnames) <= 1:
        raise ExcelTemplateError(
            f"Template must contain at least one output sheet besides '{MAPPING_SHEET_NAME}'"
        )

    workbook.remove(workbook[MAPPING_SHEET_NAME])
